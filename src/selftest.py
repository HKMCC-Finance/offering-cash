"""Self-test for the offering pipeline. Run before and after any change:

    python src/selftest.py

Covers only logic that does not need the counting machine, the scanner or the
OCR models, so it is safe to run on any laptop. Anything it cannot cover is
listed at the end as needing a real test at the church.
"""

import os
import shutil
import sys
import tempfile
import time

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.pagebreak import Break

from cash_data import (align_to_denominations, clean_denomination_frame,
                       load_offering_snapshot, save_offering_snapshot,
                       snapshot_path, subtract_offering, write_offering_block)
from check_fields import parse_courtesy_amount, parse_legal_amount, reconcile_amount
from check_scan import (CheckRecord, list_check_images, load_roster, match_roster,
                        parse_check_number, write_report, CHECK_START_ROW)
from check_summary import build_check_summary, summary_totals
from report_layout import apply_print_layout

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE = os.path.join(REPO_ROOT, "Cash_Table_Formatter.xlsx")

_passed = 0
_failed = []


def check(name, condition, detail=""):
    global _passed
    if condition:
        _passed += 1
        print(f"  PASS  {name}")
    else:
        _failed.append(name)
        print(f"  FAIL  {name} {detail}")


def test_amount_parsing():
    print("\n[수표] amount parsing")
    check("legal: words + cents", parse_legal_amount("One Hundred Twenty Five and 50/100 Dollars") == 125.50)
    check("legal: no/100", parse_legal_amount("Twenty and no/100 Dollars") == 20.00)
    check("legal: hyphenated", parse_legal_amount("Twenty-Five and 00/100") == 25.00)
    check("legal: thousands", parse_legal_amount("One Thousand Two Hundred Fifty and xx/100") == 1250.00)
    check("legal: rejects noise", parse_legal_amount("scribble nonsense") is None)
    check("courtesy: $ and commas", parse_courtesy_amount("$1,250.00") == 1250.00)
    check("courtesy: space cents", parse_courtesy_amount("$ 20 00") == 20.00)
    check("courtesy: O/0 confusion", parse_courtesy_amount("$1OO.OO") == 100.00)
    check("courtesy: rejects noise", parse_courtesy_amount("garbage") is None)


def test_reconciliation():
    print("\n[수표] courtesy vs legal reconciliation")
    check("agreement is not flagged", reconcile_amount(125.0, 125.0) == (125.0, "agree", False))
    check("disagreement is flagged", reconcile_amount(125.0, 120.0)[2] is True)
    check("single read is flagged", reconcile_amount(75.0, None)[2] is True)
    check("no read is flagged", reconcile_amount(None, None) == (None, "unreadable", True))


def test_scan_order():
    print("\n[수표] row order follows the scan, not the check number")
    directory = tempfile.mkdtemp()
    stack = ["DEP_1043_Front.tif", "DEP_207_Front.tif", "DEP_9_Front.tif",
             "DEP_1102_Front.tif", "DEP_88_Front.tif"]
    base = time.time() - 1000
    for index, name in enumerate(stack):
        path = os.path.join(directory, name)
        open(path, "wb").write(b"x")
        os.utime(path, (base + index * 10, base + index * 10))
    open(os.path.join(directory, "DEP_1043_Back.tif"), "wb").write(b"x")

    ordered = [os.path.basename(p) for p in list_check_images(directory, order="scan")]
    check("scan order preserved", ordered == stack, f"got {ordered}")
    check("back sides excluded", all(p.endswith("Front.tif") for p in ordered))
    natural = [os.path.basename(p) for p in list_check_images(directory, order="filename")]
    check("natural sort: 9 < 88 < 207 < 1043",
          natural[:3] == ["DEP_9_Front.tif", "DEP_88_Front.tif", "DEP_207_Front.tif"])
    check("check number parsed", parse_check_number("DEP_1043_Front.tif") == 1043)
    check("odd filename does not crash", parse_check_number("Front.tif") is None)


def test_roster_matching():
    print("\n[수표] donor roster matching")
    directory = tempfile.mkdtemp()
    roster_file = os.path.join(directory, "roster.csv")
    open(roster_file, "w", encoding="utf-8").write(
        "name,parish_id\nJohn Smith,101\nMary Kim,102\nPaul Lee,103\n")
    roster = load_roster(roster_file)
    check("roster loads", roster == ["John Smith", "Mary Kim", "Paul Lee"])
    check("l/1 confusion corrected", match_roster("John Smlth", roster)[0] == "John Smith")
    check("case/spacing tolerated", match_roster("MARY  KlM", roster)[0] == "Mary Kim")
    check("unknown name not forced", match_roster("Xyzzy Nobody", roster)[2] is False)


def test_check_summary():
    print("\n[수표] 수표정리 summary (#4)")
    rows = build_check_summary([20, 20, 50, 100, 5, 37.50, 150, 10, None, 25, 37.50])
    predefined = [r.amount for r in rows if r.predefined]
    extras = [r.amount for r in rows if not r.predefined]
    check("predefined rows always present", predefined == [5, 10, 20, 25, 50, 100])
    check("extras appended after $100 ascending", extras == [37.5, 150])
    check("counts correct", next(r.count for r in rows if r.amount == 20) == 2)
    check("unreadable amounts excluded", summary_totals(rows)[0] == 10)


def test_print_layout():
    print("\n[출력] print layout (#5)")
    if not os.path.exists(TEMPLATE):
        check("template present", False, f"missing {TEMPLATE}")
        return
    workbook = load_workbook(TEMPLATE)
    sheet = workbook.active
    before = sum((sheet.row_dimensions[r].height or 15.0) for r in range(1, sheet.max_row + 1))
    info = apply_print_layout(sheet)
    after = sum((sheet.row_dimensions[r].height or 15.0) for r in range(1, sheet.max_row + 1))
    check("row heights grown to fill page", after > before, f"{before} -> {after}")
    check("margins reduced", sheet.page_margins.top == 0.25 and sheet.page_margins.bottom == 0.25)
    check("fitToPage enabled", sheet.sheet_properties.pageSetUpPr.fitToPage is True)
    apply_print_layout(sheet)
    again = sum((sheet.row_dimensions[r].height or 15.0) for r in range(1, sheet.max_row + 1))
    check("idempotent on re-run", abs(again - after) < 0.5)

    # The production template prints $B$1:$L$32 on purpose: column A holds a
    # duplicate of the 수표정리 amounts in column B. Widening the range to A
    # would print that second column of numbers.
    workbook2 = load_workbook(TEMPLATE)
    sheet2 = workbook2.active
    sheet2.print_area = "B1:L32"
    apply_print_layout(sheet2)
    check("existing print area preserved", sheet2.print_area == "'Sheet1'!$B$1:$L$32",
          f"got {sheet2.print_area}")

    workbook3 = load_workbook(TEMPLATE)
    sheet3 = workbook3.active
    sheet3.print_area = None
    apply_print_layout(sheet3)
    check("print area still set when absent", bool(sheet3.print_area),
          f"got {sheet3.print_area}")

    # The production template breaks after column H on purpose: the cash and
    # check summary print on page 1, the check listing on page 2. fitToWidth=1
    # overrides a manual break and squashes both onto one sheet.
    workbook4 = load_workbook(TEMPLATE)
    sheet4 = workbook4.active
    sheet4.print_area = "B1:L32"
    sheet4.col_breaks.append(Break(id=8))
    info4 = apply_print_layout(sheet4)
    breaks4 = [b.id for b in sheet4.col_breaks.brk] if sheet4.col_breaks.brk else []
    check("manual column break preserved", breaks4 == [8], f"got {breaks4}")
    check("width not collapsed to one page",
          sheet4.sheet_properties.pageSetUpPr.fitToPage is False
          and sheet4.page_setup.fitToWidth == 0,
          f"fitToPage={sheet4.sheet_properties.pageSetUpPr.fitToPage} "
          f"fitToWidth={sheet4.page_setup.fitToWidth}")
    check("explicit print scale set for a two-page sheet",
          info4.get("pages_wide") == 2 and 10 <= info4.get("print_scale", 0) <= 100,
          f"got {info4}")
    check("scale beats the template's 68%", info4.get("print_scale", 0) > 68,
          f"got {info4.get('print_scale')}")

    # The invariant that actually guarantees the page count does not change:
    # the fraction of the usable width the content consumes must not grow.
    # Measuring columns in inches would depend on the workbook's default font,
    # so the scale is re-derived from the sheet's own previous setting instead.
    workbook5 = load_workbook(TEMPLATE)
    sheet5 = workbook5.active
    sheet5.print_area = "B1:L32"
    sheet5.col_breaks.append(Break(id=8))
    sheet5.page_setup.scale = 68
    sheet5.page_margins.left = sheet5.page_margins.right = 1.0
    before_ratio = 0.68 / (8.5 - 1.0 - 1.0)
    info5 = apply_print_layout(sheet5)
    after_ratio = (sheet5.page_setup.scale / 100.0) / (
        8.5 - sheet5.page_margins.left - sheet5.page_margins.right)
    check("column pagination cannot get worse", after_ratio <= before_ratio + 1e-9,
          f"before {before_ratio:.4f} -> after {after_ratio:.4f}")
    check("but the report does get bigger on paper",
          sheet5.page_setup.scale > 68, f"got {sheet5.page_setup.scale}")


def test_cash_pipeline():
    print("\n[현금] parsing, alignment and 2차 subtraction (#1)")
    raw = pd.DataFrame({"DENO": ["100", "50", "20", "10", "5", "2", "1", "TOTAL"],
                        "QTY": ["3", "1", "38", "50", "87", "0", "175", "354"],
                        "AMT": ["300", "50", "760", "500", "435", "0", "175", "2,220"]})
    clean = clean_denomination_frame(raw)
    check("TOTAL row dropped by type", len(clean) == 7)
    check("commas tolerated", clean["DENO"].tolist() == [1, 2, 5, 10, 20, 50, 100])

    partial = pd.DataFrame({"DENO": [1, 5, 10, 20, 100], "QTY": [10, 4, 2, 1, 1],
                            "AMT": [10, 20, 20, 20, 100]})
    aligned = align_to_denominations(partial)
    check("missing denominations re-inserted", aligned["DENO"].tolist() == [1, 2, 5, 10, 20, 50, 100])
    check("values stay with their labels",
          aligned.loc[aligned["DENO"] == 100, "QTY"].item() == 1)

    first = pd.DataFrame({"DENO": [1, 2, 5, 10, 20, 50, 100], "QTY": [175, 0, 87, 50, 38, 1, 0],
                          "AMT": [175, 0, 435, 500, 760, 50, 0]})
    second = pd.DataFrame({"DENO": [1, 2, 5, 10, 20, 50, 100], "QTY": [200, 0, 95, 60, 44, 2, 1],
                           "AMT": [200, 0, 475, 600, 880, 100, 100]})
    result = subtract_offering(second, first)
    check("2차 = 차액", result.frame["QTY"].tolist() == [25, 0, 8, 10, 6, 1, 1])
    check("clean subtraction has no warnings", not result.warnings)

    reset = pd.DataFrame({"DENO": [1, 2, 5, 10, 20, 50, 100], "QTY": [25, 0, 8, 10, 6, 1, 1],
                          "AMT": [25, 0, 40, 100, 120, 50, 100]})
    check("machine reset detected", subtract_offering(reset, first).looks_wrong)
    check("duplicate PDF detected", bool(subtract_offering(first.copy(), first).warnings))

    directory = tempfile.mkdtemp()
    path = snapshot_path(directory, "08-21-2026", "9시")
    save_offering_snapshot(path, first, pdf_file="first.pdf")
    loaded, meta = load_offering_snapshot(path)
    check("snapshot round-trips", loaded.equals(first[["DENO", "QTY", "AMT"]]))
    check("snapshot records source pdf", meta.get("source_pdf") == "first.pdf")
    check("missing snapshot is not fatal", load_offering_snapshot("/nope.json") == (None, {}))


def test_report_writing():
    print("\n[보고서] workbook writing")
    if not os.path.exists(TEMPLATE):
        check("template present", False)
        return
    directory = tempfile.mkdtemp()
    report = os.path.join(directory, "report.xlsx")
    shutil.copy(TEMPLATE, report)

    frame = align_to_denominations(pd.DataFrame(
        {"DENO": [1, 2, 5, 10, 20, 50, 100], "QTY": [175, 0, 87, 50, 38, 1, 0],
         "AMT": [175, 0, 435, 500, 760, 50, 0]}))
    workbook = load_workbook(report)
    write_offering_block(workbook.active, frame, qty_col=3, start_row=7)
    workbook.save(report)

    records = [
        CheckRecord(1, "a.tif", 1043, "John Smith", amount=100.0, amount_status="agree"),
        CheckRecord(2, "b.tif", 207, "Mary Kim", amount=20.0, amount_status="mismatch",
                    needs_review=True),
    ]
    write_report(records, report, summary_anchor=(20, 9))

    sheet = load_workbook(report).active
    check("cash block written", [sheet.cell(row=7 + i, column=3).value for i in range(7)]
          == [175, 0, 87, 50, 38, 1, 0])
    check("check block in scan order",
          [sheet.cell(row=CHECK_START_ROW + i, column=10).value for i in range(2)] == [1043, 207])
    check("review row shaded",
          sheet.cell(row=CHECK_START_ROW + 1, column=9).fill.start_color.rgb == "00FFF2CC")
    check("clean row not shaded",
          sheet.cell(row=CHECK_START_ROW, column=9).fill.start_color.rgb != "00FFF2CC")
    check("template formulas preserved", sheet["B14"].value == "=SUM(B7:B13)")
    check("summary written", sheet.cell(row=20, column=9).value == 5)


NEEDS_REAL_TEST = [
    "1차/2차 실제 카운팅 (BC-40 누적 여부 확인) - 두 PDF 모두 보관할 것",
    "스캐너 파일명 형식과 저장 시각이 실제 스캔 순서와 일치하는지",
    "check_rois.json 영역이 실제 스캔 이미지에 맞는지",
    "OCR 모델 비교 (benchmark_ocr.py, 라벨링된 수표 100~200장 필요)",
    "실제 프린터 출력 여백 확인",
    "production 양식(헌금보고서_양식.xlsx)의 열 위치 확인",
]


def main():
    print("=" * 68)
    print("Offering pipeline self-test (no machine / scanner / OCR required)")
    print("=" * 68)
    for test in (test_amount_parsing, test_reconciliation, test_scan_order,
                 test_roster_matching, test_check_summary, test_print_layout,
                 test_cash_pipeline, test_report_writing):
        test()

    print("\n" + "=" * 68)
    print(f"{_passed} passed, {len(_failed)} failed")
    if _failed:
        for name in _failed:
            print(f"  FAILED: {name}")
    print("\n교회에서 실제 테스트가 필요한 항목:")
    for item in NEEDS_REAL_TEST:
        print(f"  - {item}")
    print("=" * 68)
    return 1 if _failed else 0


if __name__ == "__main__":
    sys.exit(main())
