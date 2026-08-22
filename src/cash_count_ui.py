"""Cash counting GUI: drives the BC-40 and fills the offering report.

Data handling lives in cash_data.py so it can be tested without a GUI or a
counting machine; this file is the tkinter shell around it.
"""

import glob
import logging
import os
import shutil
import subprocess
import sys
import time
import tkinter as tk
import traceback
from datetime import datetime
from tkinter import messagebox

import pyautogui
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from cash_data import (
    align_to_denominations,
    clean_denomination_frame,
    load_offering_snapshot,
    pdf_to_dataframe,
    save_offering_snapshot,
    snapshot_path,
    subtract_offering,
    write_offering_block,
)
from report_layout import apply_print_layout

# --------------------------------------------------------------------------
# Paths and machine settings
# --------------------------------------------------------------------------
APP_PATH = r"E:\CashCounting\BC-40 UpperMonitor v13\Release\UpperMonitor.exe"
DATA_FOLDER = r"E:\CashCounting\BC-40 UpperMonitor v13\Release\Data"
REPORT_ROOT = r"E:\헌금보고서"
TEMPLATE_PATH = r"E:\헌금보고서\헌금보고서_양식.xlsx"

# Delay after launching UpperMonitor before the scripted clicks start. The
# clicks are blind, so if the window is not up yet they land on whatever is
# underneath. Raise this if the configuration steps are unreliable.
APP_LAUNCH_DELAY = 0.5
CLICK_DELAY = 0.25

# NOTE: these offsets are inherited from the original code and have NOT been
# verified against the production template (E:\헌금보고서\헌금보고서_양식.xlsx).
# Confirm before changing: the repo's Cash_Table_Formatter.xlsx puts the
# denomination labels in column A, which would make these one column too far
# right. Everything else in this file is layout-independent.
DATA_START_ROW = 7          # first denomination row
DATE_CELL = (3, 4)          # (row, column) for the mass date
TIME_CELL = (3, 6)          # (row, column) for the mass time
FIRST_OFFERING_COL = 3      # QTY column for 1차 (Amount goes in the next one)
SECOND_OFFERING_COL = 5     # QTY column for 2차

APPLY_PRINT_LAYOUT = True   # shrink margins / grow rows so printing fills the page

logger = logging.getLogger("cash_count")


def setup_logging():
    """Log to a file next to the reports.

    The packaged exe is built with console=False, so without this an exception
    is completely invisible - the window simply disappears.
    """
    try:
        os.makedirs(REPORT_ROOT, exist_ok=True)
        logging.basicConfig(
            filename=os.path.join(REPORT_ROOT, "cash_count.log"),
            level=logging.INFO,
            format="%(asctime)s %(levelname)s %(message)s",
            encoding="utf-8",
        )
    except OSError:
        logging.basicConfig(level=logging.INFO)


def report_error(title, message, exc=None):
    """Show the user an error and record the traceback."""
    if exc is not None:
        logger.error("%s: %s\n%s", title, message, traceback.format_exc())
    else:
        logger.error("%s: %s", title, message)
    messagebox.showerror(title, message)


def find_latest_pdf(data_folder, date_str):
    """Find the most recent PDF in the day's folder."""
    directory_path = os.path.join(data_folder, date_str)
    if not os.path.isdir(directory_path):
        report_error("Error", f"카운팅 데이터 폴더를 찾을 수 없습니다:\n{directory_path}")
        return None
    pdf_files = glob.glob(os.path.join(directory_path, "*.pdf"))
    if not pdf_files:
        report_error("Error", f"PDF 파일을 찾을 수 없습니다:\n{directory_path}")
        return None
    return max(pdf_files, key=os.path.getctime)


def process_pdf(pdf_file, output_dir, is_second_offering=False, mass_time=None,
                populate_header=False, previous_df=None):
    """Write one offering's counts into the report. Returns (frame, warnings)."""
    frame = align_to_denominations(clean_denomination_frame(pdf_to_dataframe(pdf_file)))
    warnings = []

    raw_df = frame.copy()
    if is_second_offering:
        if previous_df is None:
            warnings.append(
                "1차 헌금 기록을 찾을 수 없어 차감하지 못했습니다. "
                "2차 금액이 누적 합계일 수 있으니 반드시 확인하십시오."
            )
        else:
            result = subtract_offering(frame, previous_df)
            warnings.extend(result.warnings)
            if result.looks_wrong:
                # Negative counts mean the cumulative assumption broke; writing
                # them would corrupt the report, so stop and let the user decide.
                return None, warnings
            frame = result.frame

    workbook = load_workbook(output_dir)
    sheet = workbook.active

    if populate_header and mass_time:
        sheet.cell(row=DATE_CELL[0], column=DATE_CELL[1],
                   value=datetime.today().strftime("%m/%d/%y"))
        sheet.cell(row=TIME_CELL[0], column=TIME_CELL[1], value=mass_time)

    qty_col = SECOND_OFFERING_COL if is_second_offering else FIRST_OFFERING_COL
    write_offering_block(sheet, frame, qty_col, DATA_START_ROW)

    if APPLY_PRINT_LAYOUT:
        apply_print_layout(sheet)

    workbook.save(output_dir)
    logger.info("Wrote %s offering to %s",
                "2차" if is_second_offering else "1차", output_dir)
    return raw_df, warnings


def open_output_directory(output_folder):
    """Open the report folder in Explorer."""
    try:
        os.startfile(output_folder)
    except AttributeError:  # not Windows
        subprocess.Popen(["open", output_folder])
    except OSError as exc:
        report_error("Error", f"폴더를 열 수 없습니다: {exc}", exc)


def prepare_report_file(output_dir, is_second_offering):
    """Make sure the workbook we are about to write into exists."""
    if is_second_offering:
        if not os.path.exists(output_dir):
            report_error("Error", "첫 번째 헌금 보고서 파일을 찾을 수 없습니다. "
                                  "먼저 1차 헌금을 처리하십시오.")
            return False
        return True

    if os.path.exists(output_dir):
        return True

    if not os.path.exists(TEMPLATE_PATH):
        # Previously this fell through and load_workbook raised on a missing
        # file, which in the windowed exe looked like the app just vanishing.
        report_error("Error", f"보고서 양식 파일을 찾을 수 없습니다:\n{TEMPLATE_PATH}")
        return False

    shutil.copy(TEMPLATE_PATH, output_dir)
    return True


def process_cash_count_data(is_second_offering=False):
    """Locate the PDF and write it into the report."""
    mass_time = mass_time_var.get()
    cash_run_date = datetime.today().strftime("%Y%m%d")
    run_date = datetime.today().strftime("%m-%d-%Y")

    pdf_file = find_latest_pdf(DATA_FOLDER, cash_run_date)
    if not pdf_file:
        return

    output_folder = os.path.join(REPORT_ROOT, run_date)
    os.makedirs(output_folder, exist_ok=True)
    output_dir = os.path.join(output_folder, f"헌금보고서_{run_date}_{mass_time}미사.xlsx")

    if not prepare_report_file(output_dir, is_second_offering):
        return

    snapshot_file = snapshot_path(output_folder, run_date, mass_time)
    previous_df = None
    if is_second_offering:
        previous_df = first_offering_state.get("frame")
        previous_pdf = first_offering_state.get("pdf")
        if previous_df is None:
            # Fall back to disk so restarting the app between offerings does
            # not silently skip the subtraction.
            previous_df, meta = load_offering_snapshot(snapshot_file)
            previous_pdf = meta.get("source_pdf") or previous_pdf
            if previous_df is not None:
                logger.info("Loaded 1차 snapshot from disk: %s", meta)
        if previous_df is not None and previous_pdf and previous_pdf == pdf_file:
            # find_latest_pdf returned the 1차 report again, so the machine
            # never produced a 2차 one. Subtracting would just yield zeros.
            report_error("Error", "2차 헌금 PDF가 1차와 동일합니다. "
                                  "기계에서 2차 카운팅 결과가 생성되었는지 확인하십시오.")
            return

    try:
        raw_df, warnings = process_pdf(pdf_file, output_dir, is_second_offering,
                                       mass_time, not is_second_offering, previous_df)
    except Exception as exc:
        report_error("Error", f"보고서 생성 중 오류가 발생했습니다:\n{exc}", exc)
        return

    if raw_df is None:
        messagebox.showerror("Error", "\n\n".join(warnings))
        return

    for warning in warnings:
        messagebox.showwarning("확인 필요", warning)

    if not is_second_offering:
        first_offering_state["frame"] = raw_df
        first_offering_state["pdf"] = pdf_file
        try:
            save_offering_snapshot(snapshot_file, raw_df, pdf_file)
        except OSError as exc:
            logger.warning("Could not save 1차 snapshot: %s", exc)

    show_success_window(is_second_offering, output_folder)


def show_success_window(is_second_offering, output_folder):
    offering_type = "2차" if is_second_offering else "1차"
    success_win = tk.Toplevel(root)
    success_win.title("Success")
    success_win.geometry("600x300")
    tk.Label(success_win,
             text=f"{offering_type} 헌금 현금 부분의 헌금보고서 생성이 완료됐습니다.",
             font=("Arial", 18, "bold"), wraplength=550).pack(pady=30)

    if not is_second_offering and has_second_offering_var.get():
        tk.Button(success_win, text="2차 헌금 카운팅 시작",
                  command=lambda: [success_win.destroy(), start_second_offering()],
                  font=("Arial", 16, "bold"), width=20, height=2,
                  bg="#4CAF50", fg="white", cursor="hand2").pack(pady=20)
    else:
        tk.Button(success_win, text="보고서 폴더 열기",
                  command=lambda: open_output_directory(output_folder),
                  font=("Arial", 16, "bold"), width=20, height=2,
                  bg="#2196F3", fg="white", cursor="hand2").pack(pady=20)


def confirm_completion(is_second_offering=False):
    """Ask the user to confirm the machine has finished counting."""
    confirm_win = tk.Toplevel(root)
    confirm_win.title("Confirmation")
    confirm_win.geometry("700x350")

    offering_type = "2차" if is_second_offering else "1차"
    tk.Label(confirm_win,
             text=f"{offering_type} 헌금 현금 카운팅 완료 후,\n"
                  "현금을 기계 하단부에서 빼시고,\n아래 '완료' 버튼을 클릭하십시오.",
             font=("Arial", 18, "bold"), justify="center").pack(pady=40)

    def on_confirm():
        confirm_win.destroy()
        messagebox.showinfo("Processing", f"{offering_type} 헌금 현금 보고서 생성을 시작합니다.")
        process_cash_count_data(is_second_offering)

    tk.Button(confirm_win, text="완료", command=on_confirm,
              font=("Arial", 18, "bold"), width=15, height=2,
              bg="#4CAF50", fg="white", cursor="hand2").pack(pady=30)


def launch_app(app_path, is_second_offering=False):
    """Launch UpperMonitor and click through its connection settings."""
    try:
        os.startfile(app_path)
        time.sleep(APP_LAUNCH_DELAY)

        pyautogui.click(x=803, y=226)          # Port dropdown
        time.sleep(CLICK_DELAY)
        pyautogui.click(x=803, y=241)          # COM3

        pyautogui.click(x=943, y=222)          # Baud rate dropdown
        time.sleep(CLICK_DELAY)
        pyautogui.click(x=934, y=244)          # 115200

        time.sleep(CLICK_DELAY)
        pyautogui.click(x=1020, y=224)         # Open

        confirm_completion(is_second_offering)
    except Exception as exc:
        report_error("Error", f"Failed to launch application: {exc}", exc)


def start_second_offering():
    messagebox.showinfo("2차 헌금", "2차 헌금 카운팅을 시작합니다. 'OK' 버튼을 클릭하십시오...")
    confirm_completion(is_second_offering=True)


def ask_second_offering():
    """Ask whether there will be a second offering."""
    win = tk.Toplevel(root)
    win.title("2차 헌금 확인")
    win.geometry("600x350")
    tk.Label(win, text="2차 헌금이 있습니까?", font=("Arial", 20, "bold")).pack(pady=50)

    def choose(has_second):
        has_second_offering_var.set(has_second)
        win.destroy()
        proceed_to_counting()

    frame = tk.Frame(win)
    frame.pack(pady=30)
    tk.Button(frame, text="예", command=lambda: choose(True),
              font=("Arial", 18, "bold"), width=12, height=2,
              bg="#4CAF50", fg="white", cursor="hand2").pack(side=tk.LEFT, padx=20)
    tk.Button(frame, text="아니오", command=lambda: choose(False),
              font=("Arial", 18, "bold"), width=12, height=2,
              bg="#f44336", fg="white", cursor="hand2").pack(side=tk.LEFT, padx=20)


def proceed_to_counting():
    mass_time = mass_time_var.get()
    status = "있습니다" if has_second_offering_var.get() else "없습니다"
    messagebox.showinfo(
        "Selection Confirmed",
        f"{mass_time} 미사가 선택됐습니다.\n2차 헌금: {status}\n\n"
        "1차 헌금 카운팅을 시작합니다. 'OK' 버튼을 클릭하십시오...",
    )
    launch_app(APP_PATH, is_second_offering=False)


def start_cash_counting():
    """Validate the selection, then start."""
    # StringVar.set(None) stores the string "None", which is truthy - so the
    # old check never fired and reports were written as '..._None미사.xlsx'.
    if not mass_time_var.get().strip():
        messagebox.showwarning("Warning", "미사 시간을 선택하십시오.")
        return
    first_offering_state.clear()
    ask_second_offering()


setup_logging()

# 1차 counts for the session, mirrored to disk by save_offering_snapshot().
first_offering_state = {}

root = tk.Tk()
root.title("Cash Counting System")
root.geometry("600x500")

tk.Label(root, text="미사 시간을 선택하십시오.", font=("Arial", 22, "bold")).pack(pady=30)

mass_time_var = tk.StringVar(value="")
has_second_offering_var = tk.BooleanVar(value=False)

for option in ["7시반", "9시", "11시", "17시"]:
    tk.Radiobutton(root, text=option, variable=mass_time_var, value=option,
                   font=("Arial", 18, "bold"), indicatoron=1,
                   cursor="hand2").pack(pady=8)

tk.Button(root, text="현금 카운팅 시작", command=start_cash_counting,
          font=("Arial", 18, "bold"), width=18, height=2,
          bg="#2196F3", fg="white", cursor="hand2").pack(pady=40)

if __name__ == "__main__":
    root.mainloop()
