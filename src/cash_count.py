
## Load necessary pacakges 

import pyautogui
import os
import time
from datetime import datetime
import glob
import pdfplumber
import pandas as pd 
from openpyxl import load_workbook
import sys

# Function to transform the PDf to a Dataframe 
def pdf_to_dataframe(pdf_path):
    """
    Extracts table data from a single-page PDF and returns it as a pandas DataFrame.

    Parameters:
    pdf_path (str): The file path of the PDF.

    Returns:
    pd.DataFrame: A DataFrame containing the extracted table data.
    """
    # Initialize an empty DataFrame
    df = pd.DataFrame()

    # Load the PDF and extract table data from the first page
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[0]  # Only single-page PDFs are expected
        table = page.extract_table()

        # Check if there is a table and convert it to DataFrame
        if table:
            df = pd.DataFrame(table[4:], columns=table[3])  # First row as header

    return df

# Step 1: Launch the Upper Monitor Application and Set Parameters
def launch_app(app_path):
    os.startfile(app_path)
    time.sleep(3)

    # Set the Port as "COM3"
    pyautogui.click(x=798, y=149)
    time.sleep(1)
    pyautogui.click(x=761, y=170)

    # Set the Baud Rate as 115200
    pyautogui.click(x=962, y=141)
    time.sleep(1)
    pyautogui.click(x=941, y=170)

    # Set the connection value to Open
    time.sleep(1)
    pyautogui.click(x=1021, y=147)


# Step 2: Wait for the PDF to be created and locate it
def find_latest_pdf(data_folder, date_str):
    directory_path = os.path.join(data_folder, date_str)

    print(f"Please wait for the cash counter to generate a PDF in {directory_path}...")
    
    # Prompt user input to proceed with PDF processing
    input("Press 'Y' and Enter once the cash has been counted successfully: ")
    
    # Find the latest PDF file
    pdf_files = glob.glob(os.path.join(directory_path, "*.pdf"))
    if not pdf_files:
        print("No PDF files found in the directory.")
        return None
    
    latest_pdf = max(pdf_files, key=os.path.getctime)  # Get the most recent PDF file
    print(f"Found the latest PDF: {latest_pdf}")
    return latest_pdf


# Step 3: Process the PDF and save as CSV 
def process_pdf(pdf_file, output_dir):
    df = pdf_to_dataframe(pdf_file)
    df = df[:-1]
    df['DENO'] = df['DENO'].astype(int)
    df['AMT'] = df['AMT'].astype(int)
    df['QTY'] = df['QTY'].astype(int)
    df = df.sort_values(by='DENO', ascending=True)

    df.rename(columns={
        'AMT': 'Amount'
    }, inplace=True)

    df.reset_index(inplace=True, drop=True)
    df.drop(columns='DENO', inplace=True)

    # Fill-in the processed data to the formatted excel file 
    workbook = load_workbook(output_dir)
    sheet = workbook.active 

    for row_idx, row in enumerate(df.itertuples(index=False), start=2): 
        for col_idx, value in enumerate(row, start=1):
            sheet.cell(row=row_idx+5, column=col_idx+2, value=value)

    workbook.save(output_dir)


# Main Function 
def main(mass_time):
    # Path to the application
    app_path = "E:\\CashCounting\\BC-40 UpperMonitor v13\\Release\\UpperMonitor.exe"  # Update with actual path
    data_folder = "E:\\CashCounting\\BC-40 UpperMonitor v13\\Release\\Data"  # Root folder where the application saves PDFs

    # Step 1: Launch the application
    launch_app(app_path)

    # Step 2: Wait for PDF generation and locate it
    cash_run_date = datetime.today().strftime("%Y%m%d")
    pdf_file = find_latest_pdf(data_folder, cash_run_date)

    # Step 3: Process the PDF and save as CSV
    run_date = datetime.today().strftime("%m-%d-%Y")
    output_folder = os.path.join("C:\\Users\\hkmcc\\Documents\\헌금보고서", run_date)
    output_dir = output_folder + '\\헌금보고서_' + run_date + '_' + mass_time + ' 미사.xlsx'
    process_pdf(pdf_file, output_dir)


# Run the script if executed directly
if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Error: Missing mass tim argument. Use: python cash_count.py <MASS_TIME>")
        sys.exit(1)
    
    mass_time = sys.argv[1]
    main(mass_time)