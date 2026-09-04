"""Print layout for the offering report.

The report is short, so on paper it sits in the top half of the page with a
wide band of white space underneath. Growing the row heights to fill the
printable area removes that band without changing any content.
"""

import re

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

POINTS_PER_INCH = 72.0
DEFAULT_ROW_HEIGHT = 15.0  # Excel's default, used when a row has no explicit height
DEFAULT_COLUMN_WIDTH = 8.43
PAPER_WIDTH_IN = 8.5  # US Letter, portrait


def _column_width_in(worksheet, column: int) -> float:
    """Approximate a column's printed width in inches.

    Excel stores width in characters of the default font; the usual pixel
    conversion is round(width * 7) + 5 at 96 DPI for Calibri 11.
    """
    width = worksheet.column_dimensions[get_column_letter(column)].width
    if not width:
        width = DEFAULT_COLUMN_WIDTH
    return (round(width * 7) + 5) / 96.0


def _print_area_columns(worksheet):
    """(first_col, last_col) of the print area, or the sheet's used range."""
    area = worksheet.print_area
    if area:
        reference = area[0] if isinstance(area, (list, tuple)) else area
        match = re.search(r"\$?([A-Z]{1,3})\$?\d+:\$?([A-Z]{1,3})\$?\d+", str(reference))
        if match:
            from openpyxl.utils import column_index_from_string
            return (column_index_from_string(match.group(1)),
                    column_index_from_string(match.group(2)))
    return worksheet.min_column or 1, worksheet.max_column or 1


def _column_page_groups(worksheet):
    """Split the printed columns at manual column breaks.

    The offering template carries a deliberate break after column H: the cash
    and check summary print on page 1, the check listing on page 2. Forcing
    fitToWidth=1 overrides that break and squashes both onto one sheet, which
    is why width is fitted by scale here instead.
    """
    first_col, last_col = _print_area_columns(worksheet)
    breaks = []
    if worksheet.col_breaks is not None and worksheet.col_breaks.brk:
        breaks = sorted(b.id for b in worksheet.col_breaks.brk
                        if first_col <= b.id < last_col)
    groups, start = [], first_col
    for boundary in breaks:
        groups.append((start, boundary))
        start = boundary + 1
    groups.append((start, last_col))
    return groups


def apply_print_layout(worksheet, *, paper_height_in: float = 11.0,
                       top_margin_in: float = 0.25, bottom_margin_in: float = 0.25,
                       left_margin_in: float = 0.25, right_margin_in: float = 0.25,
                       header_in: float = 0.0, footer_in: float = 0.0,
                       first_row: int = 1, last_row: int = None,
                       min_row_height: float = DEFAULT_ROW_HEIGHT,
                       max_row_height: float = 45.0,
                       fill_ratio: float = 1.0,
                       vertical_center: bool = True,
                       set_print_area: bool = True) -> dict:
    """Shrink page margins and grow row heights so the report fills the page.

    Returns a small dict describing what it did, which makes the effect
    testable without opening Excel.
    """
    last_row = last_row or worksheet.max_row
    if last_row < first_row:
        return {"rows_scaled": 0, "scale": 1.0, "available_points": 0.0}

    # Capture the incoming page setup before overwriting it: the template's
    # own scale is the only reliable statement of how wide the content really
    # is. Column-width-to-inches conversion depends on the workbook's default
    # font (this template uses Aptos Narrow, not Calibri), so measuring the
    # columns directly gets the width wrong and spills onto an extra page.
    previous_scale = worksheet.page_setup.scale or 100
    previous_usable_width_in = (PAPER_WIDTH_IN
                                - (worksheet.page_margins.left or 0.0)
                                - (worksheet.page_margins.right or 0.0))

    worksheet.page_margins.top = top_margin_in
    worksheet.page_margins.bottom = bottom_margin_in
    worksheet.page_margins.left = left_margin_in
    worksheet.page_margins.right = right_margin_in
    worksheet.page_margins.header = header_in
    worksheet.page_margins.footer = footer_in

    worksheet.page_setup.orientation = "portrait"
    worksheet.print_options.horizontalCentered = True
    # Equal margins alone do not centre anything: Excel pins the block to the
    # top margin and lets every bit of leftover height fall below it. On the
    # offering report that put 0.25" above the table and 2.02" under it, which
    # reads as a bottom margin eight times the top one. Centring splits the
    # slack evenly instead.
    worksheet.print_options.verticalCentered = vertical_center

    groups = _column_page_groups(worksheet)
    has_manual_breaks = len(groups) > 1
    if has_manual_breaks:
        # Respect the manual column breaks: scale explicitly so the widest
        # page fits across, and let the width run to as many pages as the
        # breaks call for. fitToPage would collapse them into one sheet.
        worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=False)
        worksheet.page_setup.fitToWidth = 0
        worksheet.page_setup.fitToHeight = 0
    else:
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 1
        # fitToWidth/fitToHeight are ignored unless this property is set too.
        worksheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

    if set_print_area and not worksheet.print_area:
        # Only set a print area when the sheet does not already define one.
        # The production template prints $B$1:$L$32 deliberately - column A
        # holds a duplicate of the 수표정리 amounts in column B, and widening
        # the range to A would put that second column of numbers on the paper.
        # Build the reference from the column index directly: cell() returns a
        # MergedCell for merged ranges, which has no column_letter.
        last_col = get_column_letter(worksheet.max_column)
        worksheet.print_area = f"A{first_row}:{last_col}{last_row}"

    usable_in = paper_height_in - top_margin_in - bottom_margin_in - header_in - footer_in
    available_points = usable_in * POINTS_PER_INCH * fill_ratio

    heights = {}
    for row_number in range(first_row, last_row + 1):
        existing = worksheet.row_dimensions[row_number].height
        heights[row_number] = existing if existing else DEFAULT_ROW_HEIGHT

    current_total = sum(heights.values())
    if current_total <= 0:
        return {"rows_scaled": 0, "scale": 1.0, "available_points": available_points}

    print_scale = None
    if has_manual_breaks:
        # Re-scale the template's own setting by how much wider the page got.
        # The old scale paginated correctly at the old margins, so holding
        # (content width / usable width) constant reproduces exactly the same
        # column pagination - without needing to know the font's metrics.
        usable_width_in = PAPER_WIDTH_IN - left_margin_in - right_margin_in
        if previous_usable_width_in > 0:
            widened = usable_width_in / previous_usable_width_in
        else:
            widened = 1.0
        candidate = min(previous_scale * widened, 100.0)
        # Do not let the taller scale push the block off the bottom either.
        content_height_in = current_total / POINTS_PER_INCH
        if content_height_in > 0:
            candidate = min(candidate, 100.0 * usable_in / content_height_in)
        print_scale = max(10.0, candidate)
        worksheet.page_setup.scale = int(print_scale)

    scale = available_points / current_total
    if scale <= 1.0:
        result = {"rows_scaled": 0, "scale": scale, "available_points": available_points}
        if print_scale is not None:
            result["print_scale"] = worksheet.page_setup.scale
            result["pages_wide"] = len(groups)
        return result

    for row_number, height in heights.items():
        scaled = max(min_row_height, min(max_row_height, height * scale))
        worksheet.row_dimensions[row_number].height = round(scaled, 2)

    result = {
        "rows_scaled": len(heights),
        "scale": round(scale, 3),
        "available_points": round(available_points, 1),
        "new_total_points": round(sum(worksheet.row_dimensions[r].height for r in heights), 1),
    }
    if print_scale is not None:
        result["print_scale"] = worksheet.page_setup.scale
        result["pages_wide"] = len(groups)
    return result
