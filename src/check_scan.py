"""Scanned-check processing: OCR the payer name and amount, fill the report.

Changes from the original single-pass version:

* Images are processed in a deterministic, explicit order (default: the order
  they were scanned) so a row in the spreadsheet corresponds to the Nth check
  in the physical stack.
* The amount is read from both the courtesy box and the legal line and the two
  are reconciled; disagreements are flagged instead of silently accepted.
* The OCR engine is pluggable, so a replacement can be benchmarked against the
  current one without rewriting this file.
* Every read is written to a review sidecar file with the raw OCR text, so a
  volunteer can check the flagged minority instead of proofreading all of them.
"""

import argparse
import csv
import difflib
import os
import re
import unicodedata
from dataclasses import dataclass, field
from typing import List, Optional, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm

from check_fields import (
    AmountReading,
    crop,
    load_rois,
    parse_courtesy_amount,
    parse_legal_amount,
    reconcile_amount,
    upscale,
)
from check_summary import DEFAULT_PREDEFINED_AMOUNTS, build_check_summary, write_check_summary
from ocr import get_backend
from report_layout import apply_print_layout

# Where the check block lives in the report. These match the original code's
# offsets; confirm them against the production template before changing.
CHECK_START_ROW = 4
# Last row of the check block. The template's 수표정리 tally reads $L$4:$L$30,
# so this is the range the report itself considers "the checks".
CHECK_END_ROW = 30
SEQ_COL = 9          # 순번
CHECK_NUM_COL = 10   # CHECK #
PAYER_COL = 11       # 발행자
AMOUNT_COL = 12      # 금액

REVIEW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


# --------------------------------------------------------------------------
# Ordering
# --------------------------------------------------------------------------

def natural_key(name: str):
    """Sort key that orders embedded numbers numerically ('_2' before '_10')."""
    return [int(part) if part.isdigit() else part.lower()
            for part in re.split(r"(\d+)", name)]


def list_check_images(directory: str, suffix: str = "Front.tif",
                      order: str = "scan") -> List[str]:
    """Return check image paths in a deterministic order.

    os.listdir() alone returns filesystem order, which on NTFS is lexicographic
    by filename. Since the scanner names files after the MICR check number, that
    produced a spreadsheet sorted by check number rather than by scan order -
    the reported problem. The default here is 'scan': oldest file first, which
    is the order the sheets physically went through the scanner.
    """
    names = [f for f in os.listdir(directory) if f.endswith(suffix)]
    paths = [os.path.join(directory, f) for f in names]

    if order == "scan":
        # Ties (same-second writes) fall back to natural filename order so the
        # result is stable rather than arbitrary.
        paths.sort(key=lambda p: (os.path.getmtime(p), natural_key(os.path.basename(p))))
    elif order == "filename":
        paths.sort(key=lambda p: natural_key(os.path.basename(p)))
    elif order == "checknum":
        paths.sort(key=lambda p: (parse_check_number(os.path.basename(p)) or 0,
                                  natural_key(os.path.basename(p))))
    else:
        raise ValueError(f"Unknown order {order!r}; expected scan, filename or checknum")
    return paths


def parse_check_number(filename: str) -> Optional[int]:
    """Pull the check number out of the scanner's filename.

    The original code assumed exactly '<prefix>_<number>_Front.tif' and would
    raise on anything else. Fall back to the last number in the name.
    """
    stem = os.path.splitext(os.path.basename(filename))[0]
    parts = stem.split("_")
    if len(parts) > 1 and parts[1].isdigit():
        return int(parts[1])
    numbers = re.findall(r"\d+", stem)
    return int(numbers[-1]) if numbers else None


# --------------------------------------------------------------------------
# Name handling
# --------------------------------------------------------------------------

# Tokens that mark a payer-block line as address / contact detail rather than a
# name. The original "starts with a digit" test alone let real cases through:
# a payer whose address line is "Los Altos, CA 94022" has no street number, so
# the whole address was kept as part of the name.
_STREET_WORDS = r"(?:ave|avenue|st|street|dr|drive|rd|road|blvd|ln|lane|ct|court|way|pl|place|cir|circle|apt|suite|ste|hall|pkwy|terrace|ter)"
_ADDRESS_PATTERNS = (
    re.compile(r"^\d"),                                  # 4779 Sutcliff Ave
    re.compile(r"[A-Z]{2}\s*\d{5}(?:-\d{4})?"),      # CA 95118 / CA 94024-5907
    re.compile(r"\d{5}(?:-\d{4})?"),                 # bare ZIP
    re.compile(r"\d{3}[-.\s]\d{3}[-.\s]\d{4}"),          # 408-309-3828
    re.compile(r"1[-.\s]?800"),                       # 1-800 banner text
    re.compile(r"www\.|\.com", re.I),                 # printed bank URLs
    re.compile(r"" + _STREET_WORDS + r"\.?$", re.I),   # ends in a street word
    re.compile(r"^\s*" + _STREET_WORDS + r"", re.I),
)


def looks_like_address(part: str) -> bool:
    """True when a payer-block fragment is address or contact detail."""
    part = (part or "").strip()
    if not part:
        return False
    return any(p.search(part) for p in _ADDRESS_PATTERNS)


def extract_address_and_names(name_text: str):
    """Split a payer block into names and address.

    Names sit at the top of the block, so everything from the first
    address-looking fragment onwards is treated as address. Trailing noise on
    an otherwise good name line (a printed date such as "Jookyung Lee 06-10")
    is trimmed rather than dropping the whole line.
    """
    parts = [p.strip() for p in name_text.split(",")]
    flags = [looks_like_address(p) for p in parts]

    # A bare city ("Los Altos") matches nothing on its own, but it always sits
    # immediately before the "ST ZIP" fragment, so mark that neighbour too.
    state_zip = re.compile(r"^[A-Z]{2}\s*\d{5}(?:-\d{4})?$")
    for i, part in enumerate(parts):
        if state_zip.match(part) and i > 0:
            flags[i - 1] = True

    names, address_parts = [], []
    found_address = False
    for part, is_addr in zip(parts, flags):
        if is_addr:
            found_address = True
        if found_address:
            address_parts.append(part)
        else:
            # strip trailing printed dates / reference numbers off a name line
            cleaned = re.sub(r"\s+\d[\d\-/.]*$", "", part).strip()
            if cleaned:
                names.append(cleaned)
    return (", ".join(names) if names else None,
            ", ".join(address_parts) if address_parts else None)


def _normalise_name(name: str) -> str:
    name = unicodedata.normalize("NFKD", name or "")
    name = re.sub(r"[^a-zA-Z가-힣\s]", " ", name)
    return re.sub(r"\s+", " ", name).strip().lower()


def load_roster(path: Optional[str]) -> List[str]:
    """Load a donor roster (one name per line, or a CSV with a name column)."""
    if not path or not os.path.exists(path):
        return []
    names = []
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(2048)
        handle.seek(0)
        if "," in sample:
            reader = csv.DictReader(handle)
            key = None
            for candidate in ("name", "Name", "발행자", "이름", "성명"):
                if reader.fieldnames and candidate in reader.fieldnames:
                    key = candidate
                    break
            if key:
                names = [row[key] for row in reader if row.get(key)]
            else:
                handle.seek(0)
                names = [line.split(",")[0].strip() for line in handle if line.strip()]
        else:
            names = [line.strip() for line in handle if line.strip()]
    return [n for n in names if n]


def match_roster(name: Optional[str], roster: Sequence[str],
                 threshold: float = 0.82):
    """Snap an OCR'd name to the closest roster entry.

    Returns (resolved_name, score, matched). A roster is the single biggest
    available lever on name accuracy: the search space collapses from "any
    string" to "one of N known donors".
    """
    if not name or not roster:
        return name, None, False
    normalised = _normalise_name(name)
    if not normalised:
        return name, None, False
    lookup = {_normalise_name(entry): entry for entry in roster}
    candidates = difflib.get_close_matches(normalised, list(lookup), n=1, cutoff=threshold)
    if not candidates:
        return name, None, False
    best = candidates[0]
    score = difflib.SequenceMatcher(None, normalised, best).ratio()
    return lookup[best], round(score, 3), True


# --------------------------------------------------------------------------
# Extraction
# --------------------------------------------------------------------------

@dataclass
class CheckRecord:
    sequence: int
    filename: str
    check_number: Optional[int]
    payer_name: Optional[str]
    payer_name_raw: str = ""
    name_confidence: Optional[float] = None
    roster_matched: bool = False
    roster_score: Optional[float] = None
    amount: Optional[float] = None
    amount_status: str = "unreadable"
    courtesy_amount: Optional[float] = None
    legal_amount: Optional[float] = None
    courtesy_text: str = ""
    legal_text: str = ""
    needs_review: bool = False
    notes: List[str] = field(default_factory=list)


def read_amount(image, backend, rois, prefer: str = "courtesy") -> AmountReading:
    """Read the courtesy box and the legal line, then reconcile them."""
    courtesy_text = legal_text = ""
    courtesy_value = legal_value = None

    if "courtesy_amount" in rois:
        try:
            region = upscale(crop(image, rois["courtesy_amount"]))
            courtesy_text = backend.read_handwriting(region).text
            courtesy_value = parse_courtesy_amount(courtesy_text)
        except Exception:
            courtesy_text, courtesy_value = "", None

    if "legal_amount" in rois:
        try:
            region = upscale(crop(image, rois["legal_amount"]))
            legal_text = backend.read_handwriting(region).text
            legal_value = parse_legal_amount(legal_text)
        except Exception:
            legal_text, legal_value = "", None

    value, status, needs_review = reconcile_amount(courtesy_value, legal_value, prefer=prefer)
    return AmountReading(value=value, courtesy=courtesy_value, legal=legal_value,
                         status=status, needs_review=needs_review,
                         courtesy_text=courtesy_text, legal_text=legal_text)


def extract_check(image_path: str, backend, rois, sequence: int,
                  roster: Sequence[str] = (), prefer_amount: str = "courtesy",
                  name_confidence_floor: float = 0.5,
                  require_agreement: bool = True) -> CheckRecord:
    """OCR a single check image into a CheckRecord."""
    import cv2  # imported here so ordering/report helpers work without OpenCV

    record = CheckRecord(sequence=sequence, filename=os.path.basename(image_path),
                         check_number=parse_check_number(image_path), payer_name=None)

    image = cv2.imread(image_path)
    if image is None:
        record.notes.append("image could not be read")
        record.needs_review = True
        return record

    # Vision-language backends read the whole check in one pass: they are told
    # what the document is, so the amount is read as an amount instead of being
    # transcribed glyph by glyph. Anything the model will not commit to comes
    # back None and stays blank for a volunteer to fill.
    whole = backend.read_check(image)
    if whole is not None:
        record.payer_name_raw = whole.get("raw", "")
        names, _ = extract_address_and_names(whole.get("payer") or "")
        record.payer_name = names
        record.amount = whole.get("amount")
        record.amount_status = "vlm" if record.amount is not None else "unreadable"
        resolved, score, matched = match_roster(record.payer_name, roster)
        record.payer_name = resolved
        record.roster_score = score
        record.roster_matched = matched
        if record.amount is None:
            record.needs_review = True
            record.notes.append("amount not read confidently")
        if not record.payer_name:
            record.needs_review = True
            record.notes.append("name not read confidently")
        elif roster and not matched:
            record.needs_review = True
            record.notes.append("no roster match")
        return record

    try:
        name_result = backend.read_printed(upscale(crop(image, rois["payer_name"])))
        record.payer_name_raw = name_result.text
        record.name_confidence = name_result.confidence
        names, _ = extract_address_and_names(name_result.text)
        record.payer_name = names
    except Exception as exc:
        record.notes.append(f"name OCR failed: {exc}")
        record.needs_review = True

    resolved, score, matched = match_roster(record.payer_name, roster)
    record.payer_name = resolved
    record.roster_score = score
    record.roster_matched = matched

    amount = read_amount(image, backend, rois, prefer=prefer_amount)
    record.amount = amount.value
    record.amount_status = amount.status
    record.courtesy_amount = amount.courtesy
    record.legal_amount = amount.legal
    record.courtesy_text = amount.courtesy_text
    record.legal_text = amount.legal_text
    if amount.needs_review:
        record.needs_review = True
        record.notes.append(f"amount {amount.status}")
    if require_agreement and amount.status != "agree":
        # Write nothing rather than a number we cannot corroborate. Both reads
        # stay on the record so the review sheet still shows what was seen.
        record.amount = None
        record.notes.append("amount withheld (no agreement)")

    if not record.payer_name:
        record.needs_review = True
        record.notes.append("name empty")
    elif roster and not matched:
        record.needs_review = True
        record.notes.append("no roster match")
    elif (record.name_confidence is not None
          and record.name_confidence < name_confidence_floor):
        record.needs_review = True
        record.notes.append(f"low name confidence {record.name_confidence:.2f}")

    return record


# --------------------------------------------------------------------------
# Output
# --------------------------------------------------------------------------

def records_to_frame(records: Sequence[CheckRecord]) -> pd.DataFrame:
    """Report-shaped frame: 순번 / CHECK # / 발행자 / 금액, in scan order."""
    return pd.DataFrame([{
        "순번": r.sequence,
        "CHECK #": r.check_number,
        "발행자": r.payer_name,
        "금액": r.amount,
    } for r in records])


def write_review_sidecar(records: Sequence[CheckRecord], path: str) -> None:
    """Write every read, raw OCR text included, for auditing and retuning."""
    frame = pd.DataFrame([{
        "순번": r.sequence,
        "파일명": r.filename,
        "CHECK #": r.check_number,
        "발행자": r.payer_name,
        "발행자_원본OCR": r.payer_name_raw,
        "이름_신뢰도": r.name_confidence,
        "명단_일치": r.roster_matched,
        "명단_점수": r.roster_score,
        "금액": r.amount,
        "금액_숫자칸": r.courtesy_amount,
        "금액_문자줄": r.legal_amount,
        "금액_상태": r.amount_status,
        "숫자칸_원본OCR": r.courtesy_text,
        "문자줄_원본OCR": r.legal_text,
        "검토필요": r.needs_review,
        "비고": "; ".join(r.notes),
    } for r in records])
    if path.lower().endswith(".csv"):
        frame.to_csv(path, index=False, encoding="utf-8-sig")
    else:
        frame.to_excel(path, index=False)


def write_report(records: Sequence[CheckRecord], report_path: str, *,
                 highlight_review: bool = True,
                 summary_anchor: Optional[tuple] = None,
                 predefined_amounts: Sequence[float] = DEFAULT_PREDEFINED_AMOUNTS,
                 print_layout: bool = False) -> None:
    """Fill the check block of the report workbook."""
    workbook = load_workbook(report_path)
    sheet = workbook.active
    columns = (SEQ_COL, CHECK_NUM_COL, PAYER_COL, AMOUNT_COL)

    # Clear the whole block first. Two bugs live here otherwise:
    #   - openpyxl's cell(value=None) is a no-op, so a field the model declined
    #     would silently keep whatever was in that cell before;
    #   - a batch smaller than the previous one would leave the tail of the old
    #     batch behind, mixing last week's checks into this week's report.
    for row in range(CHECK_START_ROW, CHECK_END_ROW + 1):
        for column in columns:
            sheet.cell(row=row, column=column).value = None
            sheet.cell(row=row, column=column).fill = PatternFill(fill_type=None)

    for offset, record in enumerate(records):
        row = CHECK_START_ROW + offset
        if row > CHECK_END_ROW:
            raise ValueError(
                f"{len(records)} checks will not fit rows "
                f"{CHECK_START_ROW}-{CHECK_END_ROW} of the report")
        # assign directly: cell(value=None) would not clear
        sheet.cell(row=row, column=SEQ_COL).value = record.sequence
        sheet.cell(row=row, column=CHECK_NUM_COL).value = record.check_number
        sheet.cell(row=row, column=PAYER_COL).value = record.payer_name
        sheet.cell(row=row, column=AMOUNT_COL).value = record.amount
        if highlight_review and record.needs_review:
            for column in columns:
                sheet.cell(row=row, column=column).fill = REVIEW_FILL

    if summary_anchor:
        rows = build_check_summary([r.amount for r in records], predefined_amounts)
        write_check_summary(sheet, rows, start_row=summary_anchor[0],
                            amount_col=summary_anchor[1])

    if print_layout:
        apply_print_layout(sheet)

    workbook.save(report_path)


def process_checks(check_directory: str, report_filename: str, *, backend,
                   order: str = "scan", roi_path: Optional[str] = None,
                   roster_path: Optional[str] = None,
                   review_path: Optional[str] = None,
                   prefer_amount: str = "courtesy",
                   require_agreement: bool = True,
                   highlight_review: bool = True,
                   summary_anchor: Optional[tuple] = None,
                   print_layout: bool = False) -> List[CheckRecord]:
    """Process every check image in a directory into the report."""
    rois = load_rois(roi_path)
    roster = load_roster(roster_path)
    paths = list_check_images(check_directory, order=order)
    if not paths:
        raise FileNotFoundError(f"No '*Front.tif' images found in {check_directory}")

    # A vision-language backend reads whole checks in batches, which keeps the
    # GPU busy instead of paying per-image latency one check at a time.
    batch_reader = getattr(backend, "read_check_batch", None)
    if batch_reader is not None:
        import cv2

        images, loaded = [], []
        for path in paths:
            image = cv2.imread(path)
            if image is not None:
                images.append(image)
                loaded.append(path)
        replies = batch_reader(images)
        records = []
        for index, (path, reply) in enumerate(zip(loaded, replies), start=1):
            records.append(_record_from_reply(path, reply, index, roster))
    else:
        records = [
            extract_check(path, backend, rois, sequence=index, roster=roster,
                          prefer_amount=prefer_amount, require_agreement=require_agreement)
            for index, path in enumerate(tqdm(paths, desc="Processing Checks", unit="file"), start=1)
        ]

    write_report(records, report_filename, highlight_review=highlight_review,
                 summary_anchor=summary_anchor, print_layout=print_layout)

    review_path = review_path or os.path.splitext(report_filename)[0] + "_검토.xlsx"
    write_review_sidecar(records, review_path)

    flagged = sum(1 for r in records if r.needs_review)
    print(f"\nProcessed {len(records)} checks; {flagged} flagged for review.")
    print(f"Review sheet: {review_path}")
    return records


def _record_from_reply(image_path: str, reply, sequence: int,
                       roster: Sequence[str] = ()) -> CheckRecord:
    """Build a CheckRecord from a one-pass whole-check read."""
    record = CheckRecord(sequence=sequence, filename=os.path.basename(image_path),
                         check_number=parse_check_number(image_path), payer_name=None)
    record.payer_name_raw = reply.get("raw", "")
    names, _ = extract_address_and_names(reply.get("payer") or "")
    resolved, score, matched = match_roster(names, roster)
    record.payer_name = resolved
    record.roster_score = score
    record.roster_matched = matched
    record.amount = reply.get("amount")
    record.amount_status = "vlm" if record.amount is not None else "unreadable"
    if record.amount is None:
        record.needs_review = True
        record.notes.append("amount not read confidently")
    if not record.payer_name:
        record.needs_review = True
        record.notes.append("name not read confidently")
    elif roster and not matched:
        record.needs_review = True
        record.notes.append("no roster match")
    return record


def _parse_anchor(value: Optional[str]):
    if not value:
        return None
    try:
        row, col = value.split(",")
        return int(row), int(col)
    except ValueError:
        raise argparse.ArgumentTypeError("--summary-anchor expects 'row,col', e.g. 20,9")


def main(argv=None):
    parser = argparse.ArgumentParser(description="Process scanned checks into the offering report.")
    parser.add_argument("--img_dir", metavar="path", required=True, help="image file directory")
    parser.add_argument("--report_file", metavar="file", required=True, help="report file name")
    parser.add_argument("--backend", default="legacy",
                        help="OCR backend: legacy, paddleocr, qwen-vl")
    parser.add_argument("--order", default="scan", choices=("scan", "filename", "checknum"),
                        help="row order in the report (default: scan order)")
    parser.add_argument("--roi-config", default=None, help="JSON file of ROI boxes")
    parser.add_argument("--roster", default=None, help="donor roster CSV/TXT for name matching")
    parser.add_argument("--review-file", default=None, help="where to write the review sheet")
    parser.add_argument("--prefer-amount", default="courtesy", choices=("courtesy", "legal"),
                        help="which read wins when the two amounts disagree")
    parser.add_argument("--summary-anchor", default=None,
                        help="'row,col' anchor for the 수표정리 summary block")
    parser.add_argument("--allow-single-read", action="store_true",
                        help="write the best available amount even when the two "
                             "reads disagree (default: leave the cell blank)")
    parser.add_argument("--no-highlight", action="store_true",
                        help="do not shade rows that need review")
    parser.add_argument("--print-layout", action="store_true",
                        help="apply the print layout after writing")
    args = parser.parse_args(argv)

    print(f"Loading OCR backend: {args.backend}")
    backend = get_backend(args.backend)
    backend.warmup()
    print("Model loading complete.")
    print(f"Image file directory: {args.img_dir}")
    print(f"Report file name: {args.report_file}")
    print(f"Row order: {args.order}")

    process_checks(
        check_directory=args.img_dir,
        report_filename=args.report_file,
        backend=backend,
        order=args.order,
        roi_path=args.roi_config,
        roster_path=args.roster,
        review_path=args.review_file,
        prefer_amount=args.prefer_amount,
        require_agreement=not args.allow_single_read,
        highlight_review=not args.no_highlight,
        summary_anchor=_parse_anchor(args.summary_anchor),
        print_layout=args.print_layout,
    )
    print("Processing and file export complete.")


if __name__ == "__main__":
    main()
